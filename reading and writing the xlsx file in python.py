import openpyxl
def retrive():
        path='E:/samplepl.xlsx'

        df=openpyxl.load_workbook(path,read_only=True)
        active=df.active
        temp_dict={}#To store the Values of the row in dictonary format
        columns=['','',]#This is to store the columns as it starts from the 2nd col. in lineNo.29 to meet that range two space added
        for i in range(2,(active.max_column+1)):#This is to store the columns present in the xl file
                columns.append(active.cell(1,i).value)
        for row in range(2,(active.max_row)+1):
                #print(active.cell(row,1).value)
                temp_row=(active.cell(row,1).value)
                sample={'Activity':'curr_activity','RowItem':temp_row}
                for column in range(2,(active.max_column)+1):
                        if active.cell(row,column).value==None:#This for the skip of None values 
                                pass
                        else:
                                temp=columns[column]
                                temp_dict[temp]=active.cell(row,column).value
                sample['Value']=temp_dict#This is for adding the values fetch from the file
                #print(row)#This is to save the file 
                #print(active.cell(row,3).value)

wb=openpyxl.Workbook()
active=wb.active

temp_content={"data":[{"id": 1,"risk_factor": "Loan","definition":"New defination","Rare":"<2 time","Possible": "2-5 time","Almost Certain": ">5time"},
{"id": 13,"risk_factor": "Loan","definition":"New defination2","Rare":"<2 time", "Possible": "2-5 time","Almost Certain": ">5time"},
			{"id": 12,"risk_factor": "Non Performing Loan ", "Rare":"<2 time", "Possible": "2-5 time","Almost Certain": ">5time"},
			{"id": 131,"risk_factor": "Non Performing Loan ", "Rare":"<2 time", "Possible": "2-5 time","Almost Certain": ">5time"}
			]}
content=temp_content['data']
print(content)

key=[]
for c in content[0].keys():
        key.append(c)

rows=[list(key),]
for row in temp_content['data']:
        value=[]
        for obj in row.values():
                value.append(obj)
        rows.append(list(value))
        

for col in rows:
        active.append(col)



wb.save(filename='E:/test.xlsx')

